"""
═══════════════════════════════════════════════════════════════════════════
 IMPORTAÇÃO INICIAL DO ESTOQUE — Sistema de Gestão de Estoque
═══════════════════════════════════════════════════════════════════════════

 Lê a planilha "Cadastro_Inicial_Estoque.xlsx" preenchida pelos funcionários
 e popula o banco de dados, criando:
   - Almoxarifados (aba Almoxarifados)
   - Categorias    (aba Categorias)
   - Itens         (aba Itens) — cada item com quantidade > 0 gera uma ENTRADA

 COMO USAR:
   1. Coloque este arquivo na RAIZ do projeto (ao lado do manage.py)
   2. Coloque a planilha preenchida na mesma pasta
   3. Rode:   python importar_estoque.py
   4. Confira o resumo no final.

 É SEGURO RODAR: usa get_or_create, então não duplica almoxarifados nem
 categorias já existentes. Mas ITENS são sempre criados — não rode duas
 vezes para a mesma planilha, senão duplica os itens.
═══════════════════════════════════════════════════════════════════════════
"""
import os
import sys
import django
from datetime import datetime
from decimal import Decimal, InvalidOperation

# ── Configurar Django ─────────────────────────────────────────────
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'almoxgeo.settings')
django.setup()

from django.db import transaction
from estoque.models import (
    Almoxarifado, Categoria, Fornecedor, Item, Movimentacao, Usuario
)

try:
    from openpyxl import load_workbook
except ImportError:
    print('ERRO: openpyxl não instalado. Rode: pip install openpyxl')
    sys.exit(1)

PLANILHA = 'Cadastro_Inicial_Estoque.xlsx'

# ── Mapas de tradução (texto da planilha → valor do banco) ────────
UNIDADES = {
    'Unidade': 'un', 'Caixa': 'cx', 'Resma': 'rs', 'Litro': 'lt',
    'Quilograma': 'kg', 'Metro': 'm', 'Pacote': 'pc', 'Outros': 'ou',
}
TIPOS_MATERIAL = {
    'Material de Consumo': 'CONSUMO',
    'Bem Permanente': 'PERMANENTE',
}
TIPOS_EMPENHO = {
    'Ordinário': 'ORDINARIO',
    'Estimativo': 'ESTIMATIVO',
    'Global': 'GLOBAL',
    '': '',
}


def limpar(valor):
    """Retorna string limpa ou '' para células vazias."""
    if valor is None:
        return ''
    return str(valor).strip()


def parse_int(valor, padrao=0):
    s = limpar(valor)
    if not s:
        return padrao
    try:
        return int(float(s.replace(',', '.')))
    except (ValueError, TypeError):
        return padrao


def parse_decimal(valor):
    """Converte '12,50' ou '12.50' em Decimal. Retorna None se vazio."""
    s = limpar(valor)
    if not s:
        return None
    s = s.replace('R$', '').replace(' ', '').replace('.', '').replace(',', '.') \
        if ',' in s else s.replace('R$', '').replace(' ', '')
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def parse_data(valor):
    """Converte data da planilha (datetime ou texto DD/MM/AAAA) em date. None se vazio."""
    if valor is None or limpar(valor) == '':
        return None
    if isinstance(valor, datetime):
        return valor.date()
    s = limpar(valor)
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def linha_vazia(valores):
    """True se a linha não tem nome de item (col 1)."""
    return not limpar(valores[0])


def eh_exemplo(nome):
    """Detecta as linhas de exemplo para pular."""
    exemplos = {
        'Caneta esferográfica azul', 'Papel A4 75g',
        'Microscópio óptico binocular', 'Álcool etílico 70% 1L',
    }
    return nome in exemplos


def main():
    if not os.path.exists(PLANILHA):
        print(f'ERRO: planilha "{PLANILHA}" não encontrada nesta pasta.')
        sys.exit(1)

    print(f'Lendo {PLANILHA}...\n')
    wb = load_workbook(PLANILHA, data_only=True)

    # ── Responsável padrão para almoxarifados novos ───────────────
    # Usa o primeiro superusuário. Ajuste se quiser outro.
    responsavel = Usuario.objects.filter(is_superuser=True).first()
    if not responsavel:
        responsavel = Usuario.objects.first()
    if not responsavel:
        print('ERRO: nenhum usuário no banco. Crie um superusuário primeiro:')
        print('       python manage.py createsuperuser')
        sys.exit(1)
    print(f'Responsável padrão dos almoxarifados: {responsavel}\n')

    resumo = {'almox': 0, 'cat': 0, 'forn': 0, 'itens': 0, 'pulados': 0, 'erros': []}

    with transaction.atomic():
        # ── 1) ALMOXARIFADOS ──────────────────────────────────────
        if 'Almoxarifados' in wb.sheetnames:
            ws = wb['Almoxarifados']
            for row in ws.iter_rows(min_row=4, values_only=True):
                nome = limpar(row[0])
                if not nome or nome in ('Almoxarifado Central', 'Almox. Lab. Geoquímica'):
                    continue
                loc = limpar(row[1]) if len(row) > 1 else ''
                desc = limpar(row[2]) if len(row) > 2 else ''
                obj, criado = Almoxarifado.objects.get_or_create(
                    nome=nome,
                    defaults={'localizacao': loc or 'A definir',
                              'descricao': desc, 'responsavel': responsavel},
                )
                if criado:
                    resumo['almox'] += 1

        # ── 2) CATEGORIAS ─────────────────────────────────────────
        if 'Categorias' in wb.sheetnames:
            ws = wb['Categorias']
            exemplos_cat = {'Material de Escritório', 'Reagentes', 'Vidraria',
                            'Informática', 'Limpeza'}
            for row in ws.iter_rows(min_row=4, values_only=True):
                nome = limpar(row[0])
                if not nome or nome in exemplos_cat:
                    continue
                desc = limpar(row[1]) if len(row) > 1 else ''
                obj, criado = Categoria.objects.get_or_create(
                    nome=nome, defaults={'descricao': desc},
                )
                if criado:
                    resumo['cat'] += 1

        # ── 3) ITENS ──────────────────────────────────────────────
        ws = wb['Itens']
        for n_linha, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
            if linha_vazia(row):
                continue
            nome = limpar(row[0])
            if eh_exemplo(nome):
                resumo['pulados'] += 1
                continue

            nome_almox = limpar(row[1])
            nome_cat = limpar(row[2])

            # Validar obrigatórios
            if not nome_almox or not nome_cat:
                resumo['erros'].append(f'Linha {n_linha}: "{nome}" sem almoxarifado ou categoria. Pulado.')
                continue

            # Buscar/criar FK (get_or_create cobre quem não pôs nas abas)
            almox, _ = Almoxarifado.objects.get_or_create(
                nome=nome_almox,
                defaults={'localizacao': 'A definir', 'responsavel': responsavel},
            )
            categoria, _ = Categoria.objects.get_or_create(nome=nome_cat)

            # Fornecedor (opcional)
            fornecedor = None
            nome_forn = limpar(row[14]) if len(row) > 14 else ''
            if nome_forn:
                fornecedor, criado_f = Fornecedor.objects.get_or_create(nome=nome_forn)
                if criado_f:
                    resumo['forn'] += 1

            quantidade = parse_int(row[3])

            # Criar item (quantidade_atual começa em 0; a ENTRADA ajusta)
            item = Item(
                nome=nome,
                almoxarifado=almox,
                categoria=categoria,
                quantidade_atual=0,
                unidade_medida=UNIDADES.get(limpar(row[4]), 'un'),
                tipo_material=TIPOS_MATERIAL.get(limpar(row[5]), 'CONSUMO'),
                quantidade_minima=parse_int(row[6]),
                localizacao_fisica=limpar(row[7]),
                descricao=limpar(row[8]),
                lote=limpar(row[9]),
                data_fabricacao=parse_data(row[10]),
                data_validade=parse_data(row[11]),
                valor=parse_decimal(row[12]),
                codigo_patrimonio=limpar(row[13]),
                fornecedor=fornecedor,
                numero_empenho=limpar(row[15]),
                tipo_empenho=TIPOS_EMPENHO.get(limpar(row[16]), ''),
                processo_sei=limpar(row[17]),
                nf_numero=limpar(row[18]),
                nf_serie=limpar(row[19]),
                nf_data_emissao=parse_data(row[20]),
                nf_data_entrada=parse_data(row[21]),
                nf_valor=parse_decimal(row[22]),
            )
            item.save()  # gera código interno
            resumo['itens'] += 1

            # Registrar ENTRADA inicial se quantidade > 0
            if quantidade > 0:
                mov = Movimentacao(
                    item=item,
                    tipo='ENTRADA',
                    quantidade=quantidade,
                    responsavel=responsavel,
                    observacao='Cadastro inicial do estoque (importação via planilha).',
                )
                mov.save()  # atualiza quantidade_atual do item

    # ── Resumo ────────────────────────────────────────────────────
    print('═' * 55)
    print(' IMPORTAÇÃO CONCLUÍDA')
    print('═' * 55)
    print(f'  Almoxarifados criados : {resumo["almox"]}')
    print(f'  Categorias criadas    : {resumo["cat"]}')
    print(f'  Fornecedores criados  : {resumo["forn"]}')
    print(f'  Itens cadastrados     : {resumo["itens"]}')
    print(f'  Linhas de exemplo puladas: {resumo["pulados"]}')
    if resumo['erros']:
        print(f'\n  ATENÇÃO — {len(resumo["erros"])} linha(s) com problema:')
        for e in resumo['erros']:
            print(f'    • {e}')
    print('═' * 55)


if __name__ == '__main__':
    main()
